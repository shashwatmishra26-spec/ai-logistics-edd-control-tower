"""
NDR Pending-Response Outreach Export (Excel).

Purpose: a single workbook, one sheet per outreach channel (IVR, WhatsApp,
Manual Agent Call, Email), listing ONLY the customers who are currently
PENDING a response on that channel — i.e. shipments still open with an
unresolved NDR event. This is the file attached to the outreach email drafted
for the responsible team.

Why "pending" needs no extra modeling: outputs/customer_care_notifications.csv
(the NDR queue) is built from `df[(df["is_open"]) & (df["has_ndr"])]` — by
construction, every row in it is a shipment whose failed-delivery event has
NOT yet been resolved (no successful reattempt, no RTO decision). A shipment
that got a customer response and either delivered or converted to RTO drops
out of this queue entirely (it becomes is_open=False). So "still in this
queue" IS "still pending a response" for this dataset — no separate
has_responded flag needs to be invented.

Each shipment is assigned exactly ONE primary channel (recommended_channel,
computed in src/ndr_agent/ndr_recovery.py::assign_ndr_channel) plus an
optional parallel WhatsApp flag (also_whatsapp) for reasons where the
customer needs to take a specific action. This is what prevents the same
customer being contacted on multiple redundant channels for the same open
case — each customer appears on exactly one "primary channel" sheet, and
additionally on the WhatsApp sheet only if also_whatsapp=True.

PRIVACY: no real name/phone/address is used anywhere in this workbook — see
the privacy note in src/ndr_agent/ndr_consolidated_report.py. Contact is
resolved at send-time via `contact_lookup_key` against a secure CRM.
"""

from datetime import datetime
from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parent.parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from config.config import (  # noqa: E402
    FEATURED_SHIPMENTS_PATH,
    NDR_CHANNEL_ROUTING_PATH,
    NDR_PENDING_RESPONSE_XLSX_PATH,
    NDR_QUEUE_PATH,
    OUTPUTS_DIR,
    SNAPSHOT_DATE,
)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10.5)
BODY_FONT = Font(name=FONT_NAME, size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SHEET_COLUMNS = [
    ("shipment_id", "Shipment ID (contact_lookup_key)", 26),
    ("awb", "AWB", 14),
    ("order_id", "Order ID", 12),
    ("carrier", "Carrier", 12),
    ("lane_class", "Lane Class", 11),
    ("ndr_reason", "NDR Reason", 22),
    ("attempt_number", "Attempt #", 9),
    ("priority", "Priority", 12),
    ("also_whatsapp", "Also on WhatsApp?", 15),
    ("deadline", "Response Deadline", 17),
    ("channel_rationale", "Why This Channel", 46),
]

CHANNEL_SHEETS = [
    ("IVR", "IVR"),
    ("WhatsApp (Parallel)", "WHATSAPP"),
    ("Manual Agent Call", "MANUAL_CALL"),
    ("Email", "EMAIL"),
]


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def _write_table(ws, df, columns):
    headers = [label for _, label, _ in columns]
    ws.append(headers)
    _style_header(ws, len(headers))
    ws.freeze_panes = "A2"
    for _, row in df.iterrows():
        vals = []
        for key, _, _ in columns:
            v = row.get(key, "")
            if key == "also_whatsapp":
                v = "Yes" if bool(v) else "No"
            vals.append(v)
        ws.append(vals)
    for i, (_, _, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(columns[c - 1][0] == "channel_rationale"))


def build_workbook(ndr_queue: pd.DataFrame, snapshot_date: str) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # --- Read Me sheet -----------------------------------------------------
    ws = wb.create_sheet("Read Me")
    ws.column_dimensions["A"].width = 108
    ws["A1"] = "NDR Pending-Response Outreach — Read Me"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Snapshot date: {snapshot_date}  |  Generated: {datetime.utcnow().strftime('%Y-%m-%d')} (UTC)"
    ws["A2"].font = SUBTITLE_FONT

    lines = [
        "",
        "WHAT THIS FILE IS",
        "Each sheet after this one lists customers on ONE outreach channel — IVR, WhatsApp, "
        "Manual Agent Call, or Email — who currently have an OPEN, unresolved failed-delivery "
        "(NDR) event and have NOT yet had their case resolved.",
        "",
        "WHY 'PENDING' NEEDS NO GUESSWORK",
        "The source queue (outputs/customer_care_notifications.csv) only ever contains shipments "
        "that are still open AND still carrying an NDR event. The moment a customer responds and "
        "the shipment either delivers or converts to RTO, it drops out of this queue entirely. So "
        "every row in every sheet below is, by construction, still pending a response — nothing "
        "here is inferred or fabricated.",
        "",
        "ONE PRIMARY CHANNEL PER CUSTOMER — NO DUPLICATE CONTACT",
        "Every shipment is assigned exactly ONE primary channel (IVR / Manual Agent Call / Email), "
        "so the same customer is never called AND texted AND emailed for the same open case. "
        "WhatsApp is the one sanctioned PARALLEL channel: it runs alongside the primary channel "
        "(never instead of it) specifically when the NDR reason requires the customer to take an "
        "action — confirm a delivery slot, share a location pin, verify an incomplete address, or "
        "confirm COD readiness. A shipment can therefore appear on its primary-channel sheet AND "
        "on the WhatsApp sheet, but never on two primary-channel sheets at once.",
        "",
        "CHANNEL ROUTING LOGIC (see config/config.py + src/ndr_agent/ndr_recovery.py)",
        "  1. Manual Agent Call (Rs 15-25/call, the expensive channel) — gated by severity: "
        "2nd/3rd+ delivery attempt, a high-value COD-payment dispute, a high-risk reason, or a "
        "case aged past ~36 hours since first attempt.",
        "  2. Email — backup/documentation channel, used specifically when the phone itself is "
        "unreachable; this is also the paper trail if a dispute needs evidence later.",
        "  3. IVR (automated call) — the default first-touch channel for simple, low-complexity "
        "reasons; near-zero cost, used for the bulk of Day-1 volume.",
        "  + WhatsApp runs in parallel whenever the customer needs to actively do something.",
        "",
        "PRIVACY",
        "No real customer name, phone number, or address appears anywhere in this workbook. Each "
        "row carries a 'Shipment ID (contact_lookup_key)' only — the actual contact details are "
        "resolved at send-time via a secure CRM lookup on that key. This mirrors the same "
        "PII-safe design used throughout the AI Logistics EDD Control Tower "
        "(see src/ndr_agent/ndr_consolidated_report.py).",
        "",
        "DATA CONFIDENCE",
        "Shipment/reason/attempt fields are ACTUAL (from the source workbook). recommended_channel "
        "and also_whatsapp are DERIVED via a deterministic, documented rule cascade (not ML, not "
        "random) — fully reproducible from outputs/customer_care_notifications.csv. This workbook "
        "itself, and any outreach content referencing it, is MOCK — no message is actually sent.",
        "",
        "SHEETS IN THIS WORKBOOK",
    ]
    for label, _ in CHANNEL_SHEETS:
        lines.append(f"  - {label}")

    r = 3
    for line in lines:
        cell = ws.cell(row=r, column=1, value=line)
        if line and line == line.upper() and not line.startswith(" "):
            cell.font = Font(name=FONT_NAME, bold=True, size=11)
        else:
            cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # --- Per-channel sheets --------------------------------------------------
    for label, key in CHANNEL_SHEETS:
        ws = wb.create_sheet(label[:31])
        if key == "WHATSAPP":
            sub = ndr_queue[ndr_queue.get("also_whatsapp", False) == True]  # noqa: E712
        else:
            sub = ndr_queue[ndr_queue["recommended_channel"] == "IVR"] if key == "IVR" else \
                  ndr_queue[ndr_queue["recommended_channel"] == "Manual Agent Call"] if key == "MANUAL_CALL" else \
                  ndr_queue[ndr_queue["recommended_channel"] == "Email"]
        priority_rank = {"P1 - Urgent": 0, "P2 - High": 1, "P3 - Standard": 2}
        sub = sub.copy()
        sub["_rank"] = sub["priority"].map(priority_rank)
        sub = sub.sort_values(["_rank", "attempt_number"], ascending=[True, False]).drop(columns="_rank")
        _write_table(ws, sub, SHEET_COLUMNS)

    return wb


def run():
    ndr_queue = pd.read_csv(NDR_QUEUE_PATH)
    if "recommended_channel" not in ndr_queue.columns:
        raise RuntimeError(
            "customer_care_notifications.csv has no recommended_channel column — "
            "run src.ndr_agent.ndr_recovery first."
        )

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

    # Standalone, documented channel-routing CSV (source of truth for the workbook).
    routing_cols = [
        "shipment_id", "awb", "order_id", "carrier", "lane_class", "ndr_reason",
        "attempt_number", "priority", "recommended_channel", "also_whatsapp",
        "channel_rationale", "deadline",
    ]
    ndr_queue[[c for c in routing_cols if c in ndr_queue.columns]].to_csv(NDR_CHANNEL_ROUTING_PATH, index=False)

    wb = build_workbook(ndr_queue, SNAPSHOT_DATE)
    wb.save(NDR_PENDING_RESPONSE_XLSX_PATH)

    print(f"Wrote channel routing -> {NDR_CHANNEL_ROUTING_PATH}")
    print(f"Wrote pending-response outreach workbook -> {NDR_PENDING_RESPONSE_XLSX_PATH}")
    for label, key in CHANNEL_SHEETS:
        if key == "WHATSAPP":
            n = int((ndr_queue.get("also_whatsapp", False) == True).sum())  # noqa: E712
        else:
            chan = {"IVR": "IVR", "MANUAL_CALL": "Manual Agent Call", "EMAIL": "Email"}[key]
            n = int((ndr_queue["recommended_channel"] == chan).sum())
        print(f"  {label}: {n} pending")
    return ndr_queue


if __name__ == "__main__":
    run()
